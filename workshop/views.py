import json
from django.utils import timezone
import csv
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Produto, Seccao, Venda, VendeSe, Loja, Owner, Fornecedor, Contacto, Evento
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction



def produtosHTMLFUNC(request):
    return render(request, 'produtosHTML.html', {})

def produtosTESTEFUNC(request):
    produtos = Produto.objects.all()
    return render(request, 'produtosTESTE.html', {'produtos': produtos})

def vendasTESTEFUNC(request):
    vendas = Venda.objects.all()
    return render(request, 'vendasTESTE.html', {'vendas': vendas})

def nova_venda(request):
    loja_nome = request.session.get('loja_atual')
    if not loja_nome:
        return redirect('home')

    if request.method == 'POST':
        try:
            with transaction.atomic():
                dados = json.loads(request.body)
                itens = dados.get('itens', [])
                metodo = dados.get('metodo_pagamento', 'dinheiro')
                contacto_id = dados.get('contacto_id')

                # Recibo gerado no servidor, de forma segura
                ultima_venda = Venda.objects.select_for_update().order_by('-recibo').first()
                recibo = (ultima_venda.recibo + 1) if ultima_venda else 1001

                total = sum(item['preco'] * item['qty'] for item in itens)

                contacto = None
                if metodo == 'mbway' and contacto_id:
                    contacto = Contacto.objects.filter(id=contacto_id).first()

                evento = Evento.objects.filter(
                    loja__nomeloja=loja_nome, ativo=True
                ).first()   

                venda = Venda.objects.create(
                    recibo=recibo,
                    total=round(total, 2),
                    data=timezone.now(),
                    metodo_pagamento=metodo,
                    contacto=contacto,
                    evento=evento,        
                )
                for item in itens:
                    produto = Produto.objects.get(id=item['id'])
                    VendeSe.objects.create(
                        venda=venda, produto=produto,
                        qtd=item['qty'], preco=item['preco']
                    )
                    produto.qtd -= item['qty']
                    produto.save()

                return JsonResponse({'sucesso': True, 'recibo': recibo})

        except Exception as e:
            import traceback
            print(f"ERRO NOVA_VENDA: {e}")
            print(traceback.format_exc())
            return JsonResponse({'erro': str(e)}, status=500)
    
    # GET
    ultima_venda = Venda.objects.order_by('-recibo').first()
    proximo_recibo = (ultima_venda.recibo + 1) if ultima_venda else 1001

    produtos_lista = []
    for p in Produto.objects.filter(loja__nomeloja=loja_nome, oculto=False).order_by('id'):
        produtos_lista.append({
            'id': p.id,
            'nome': p.nome,
            'preco': float(p.preco),
            'qtd': p.qtd,
            'desconto': p.desconto if p.desconto else 0,
            'iva': p.iva if p.iva else 0,
            'seccao': str(p.seccao_id),
            'fornecedor': p.fornecedor.nome if p.fornecedor else None,
        })

    contactos = list(Contacto.objects.values('id', 'nome', 'telefone').order_by('nome'))

    return render(request, 'nova_venda.html', {
        'produtos_json': json.dumps(produtos_lista),
        'contactos_json': json.dumps(contactos),
        'proximo_recibo': proximo_recibo,
        'loja_atual': loja_nome,
    })


@login_required(login_url='login')
def criar_contacto(request):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
            nome = dados.get('nome', '').strip()
            telefone = dados.get('telefone', '').strip()
            if not nome or not telefone:
                return JsonResponse({'erro': 'Nome e telefone são obrigatórios.'}, status=400)
            contacto = Contacto.objects.create(nome=nome, telefone=telefone)
            return JsonResponse({'sucesso': True, 'id': contacto.id, 'nome': contacto.nome, 'telefone': contacto.telefone})
        except Exception as e:
            return JsonResponse({'erro': str(e)}, status=400)
    return JsonResponse({'erro': 'Método não permitido'}, status=405)


@login_required(login_url='login')
def home(request):
    lojas = Loja.objects.select_related('owner').filter(oculta=False)
    owners = Owner.objects.all() 
    loja_nome = request.session.get('loja_atual')
    loja_sessao = Loja.objects.select_related('owner').filter(nomeloja=loja_nome).first() if loja_nome else None
    return render(request, 'home.html', {'lojas': lojas, 'owners': owners, 'loja_sessao': loja_sessao})

def trocar_loja(request):
    erro = ''
    if request.method == 'POST':
        password = request.POST.get('password')
        user = authenticate(request, username=request.user.username, password=password)
        if user:
            request.session.pop('loja_atual', None)
            return redirect('home')
        else:
            erro = 'Password incorreta.'
    return render(request, 'trocar_loja.html', {'erro': erro})

def login_view(request):
    erro = ''
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('lista_vendas')
        else:
            erro = 'Utilizador ou password incorretos.'
    return render(request, 'login.html', {'erro': erro})

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required(login_url='login')
def lista_vendas(request):
    loja_nome = request.session.get('loja_atual')
    if not loja_nome:
        return redirect('home')

    data_filtro   = request.GET.get('data', '')
    evento_filtro = request.GET.get('evento', '')

    # Se não veio filtro de evento, usa o evento ativo por defeito
    if not evento_filtro:
        evento_ativo = Evento.objects.filter(loja__nomeloja=loja_nome, ativo=True).first()
        if evento_ativo:
            evento_filtro = str(evento_ativo.id)

    vendas = Venda.objects.filter(
        linhas__produto__seccao__loja__nomeloja=loja_nome
    ).distinct().order_by('-data')

    if data_filtro:
        vendas = vendas.filter(data__date=data_filtro)

    if evento_filtro:
        vendas = vendas.filter(evento__id=evento_filtro)

    eventos = Evento.objects.filter(loja__nomeloja=loja_nome).order_by('-data_inicio')

    return render(request, 'lista_vendas.html', {
        'vendas':        vendas,
        'data_filtro':   data_filtro,
        'evento_filtro': evento_filtro,
        'eventos':       eventos,
        'loja_atual':    loja_nome,
    })

# para exportar as vendas para CSV

@login_required(login_url='login')
def exportar_vendas(request):
    loja_nome     = request.session.get('loja_atual')
    evento_filtro = request.GET.get('evento', '')

    wb = Workbook()
    wb.remove(wb.active)

    cabecalho = [
        'Recibo', 'Data', 'Loja', 'Evento',
        'Método Pagamento', 'Recebedor', 'Total (€)',
        'Produto', 'Quantidade', 'Preço Unit. (€)', 'Subtotal (€)',
    ]

    vendas = Venda.objects.prefetch_related(
        'linhas__produto__seccao__loja',
        'contacto', 'evento',
    ).order_by('-data')

    if loja_nome:
        vendas = vendas.filter(
            linhas__produto__seccao__loja__nomeloja=loja_nome
        ).distinct()

    if evento_filtro:
        vendas = vendas.filter(evento__id=evento_filtro)

    folhas = {}

    for venda in vendas:
        recebedor   = '—'
        nome_evento = venda.evento.nome if venda.evento else '—'
        if venda.metodo_pagamento == 'mbway' and venda.contacto:
            recebedor = f'{venda.contacto.nome} ({venda.contacto.telefone})'

        linhas = list(venda.linhas.all())

        for i, linha in enumerate(linhas):
            loja      = linha.produto.seccao.loja
            nome_loja = loja.nomeloja

            if nome_loja not in folhas:
                ws = wb.create_sheet(title=nome_loja[:31])
                ws.append(cabecalho)
                for col in range(1, len(cabecalho) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font      = Font(bold=True)
                    cell.fill      = PatternFill(start_color='C8F060', end_color='C8F060', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                folhas[nome_loja] = ws

            ws = folhas[nome_loja]

            # Só a primeira linha da venda tem os dados do cabeçalho da venda
            if i == 0:
                campos_venda = [
                    venda.recibo,
                    venda.data.strftime('%d/%m/%Y %H:%M'),
                    nome_loja,
                    nome_evento,
                    venda.metodo_pagamento,
                    recebedor,
                    float(venda.total),
                ]
            else:
                campos_venda = ['', '', '', '', '', '', '']

            ws.append(campos_venda + [
                linha.produto.nome,
                linha.qtd,
                float(linha.preco),
                round(float(linha.preco) * linha.qtd, 2),
            ])

    larguras = [10, 18, 20, 18, 18, 28, 12, 25, 12, 16, 14]
    for ws in folhas.values():
        for i, largura in enumerate(larguras, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = largura

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="vendas_por_loja.xlsx"'
    wb.save(response)
    return response

## apagar dados

@login_required(login_url='login')
def limpar_vendas(request):
    erro = ''
    if request.method == 'POST':
        password = request.POST.get('password')
        user = authenticate(request, username=request.user.username, password=password)
        if user:
            VendeSe.objects.all().delete()
            Venda.objects.all().delete()
            return redirect('lista_vendas')
        else:
            erro = 'Password incorreta.'
    return render(request, 'limpar_vendas.html', {'erro': erro})


def lista_lojas(request):
    lojas = Loja.objects.select_related('owner').filter(oculta=False)
    return render(request, 'lista_lojas.html', {'lojas': lojas})

def lista_owners(request):
    owners = Owner.objects.prefetch_related('lojas').all()
    return render(request, 'lista_owners.html', {'owners': owners})

def selecionar_loja(request):
    if request.method == 'POST':
        loja_nome = request.POST.get('loja')
        if Loja.objects.filter(nomeloja=loja_nome, oculta=False).exists():        
            request.session['loja_atual'] = loja_nome
    return redirect('home')


@login_required(login_url='login')
def adicionar_produto(request):
    loja_nome = request.session.get('loja_atual')
    if not loja_nome:
        return redirect('home')

    loja = Loja.objects.get(nomeloja=loja_nome)
    seccoes = Seccao.objects.filter(loja__nomeloja=loja_nome)
    fornecedores = Fornecedor.objects.all().order_by('nome')
    erro = ''
    sucesso = ''
    if request.GET.get('sucesso'):
        sucesso = f'Produto "{request.GET["sucesso"]}" adicionado com sucesso!'

    if request.method == 'POST':
        try:
            nome           = request.POST.get('nome', '').strip()
            preco          = float(request.POST.get('preco'))  
            qtd            = int(request.POST.get('qtd'))
            desconto       = int(request.POST.get('desconto', 0))
            iva            = int(request.POST.get('iva', 0))
            seccao_nome    = request.POST.get('seccao')
            fornecedor_nif = request.POST.get('fornecedor') or None

            if not nome:
                erro = 'O nome não pode estar vazio.'
            else:
                seccao = Seccao.objects.get(nome=seccao_nome)
                fornecedor = Fornecedor.objects.get(nif=fornecedor_nif) if fornecedor_nif else None
                Produto.objects.create(
                    nome=nome,
                    preco=preco,
                    qtd=qtd,
                    desconto=desconto,
                    iva=iva,
                    seccao=seccao,
                    loja=loja,
                    fornecedor=fornecedor,
                )
                return redirect(f'/adicionar_produto/?sucesso={nome}')
        except Exception as e:
            erro = f'Erro: {e}'

    return render(request, 'adicionar_produto.html', {
        'seccoes': seccoes,
        'fornecedores': fornecedores,
        'loja_atual': loja_nome,
        'erro': erro,
        'sucesso': sucesso,
    })


@login_required(login_url='login')
def adicionar_seccao(request):
    loja_nome = request.session.get('loja_atual')
    if not loja_nome:
        return redirect('home')

    erro = ''
    sucesso = ''

    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if not nome:
            erro = 'O nome não pode estar vazio.'
        elif Seccao.objects.filter(nome=nome).exists():
            erro = f'Já existe uma secção com o nome "{nome}".'
        else:
            loja = Loja.objects.get(nomeloja=loja_nome)
            Seccao.objects.create(nome=nome, loja=loja)
            sucesso = f'Secção "{nome}" criada com sucesso!'

    return render(request, 'adicionar_seccao.html', {
        'loja_atual': loja_nome,
        'erro': erro,
        'sucesso': sucesso,
    })



def remover_loja(request, nomeloja):
    loja = get_object_or_404(Loja, nomeloja=nomeloja)
    if request.method == 'POST':
        password = request.POST.get('password')
        user = authenticate(username=request.user.username, password=password)
        if user is not None:
            produtos = Produto.objects.filter(loja=loja)
            VendeSe.objects.filter(produto__in=produtos).delete()
            produtos.delete()
            Seccao.objects.filter(loja=loja).delete()
            loja.delete()
        else:
            from django.contrib import messages
            messages.error(request, f'Password incorreta. A loja "{nomeloja}" não foi removida.')
    return redirect('home')

def adicionar_loja(request):
    if request.method == 'POST':
        cc = request.POST.get('cc')
        # se veio um cc novo, cria o owner primeiro
        if cc:
            Owner.objects.get_or_create(
                cc=cc,
                defaults={
                    'name': request.POST['name'],
                    'birth': request.POST['birth'],
                    'email': request.POST['owner_email'],
                    'contact': request.POST['contact'],
                    'morada': request.POST.get('morada', False) == 'on'
                }
            )
        owner = get_object_or_404(Owner, cc=request.POST['owner_cc'] if not cc else cc)
        Loja.objects.create(
            nomeloja=request.POST['nomeloja'],
            localizacao=request.POST['localizacao'],
            capitalsocial=request.POST['capitalsocial'],
            emailloja=request.POST['emailloja'],
            contacto=request.POST['contacto'],
            owner=owner
        )
    return redirect('home')

def adicionar_owner(request):
    if request.method == 'POST':
        Owner.objects.create(
            cc=request.POST['cc'],
            name=request.POST['name'],
            birth=request.POST['birth'],
            email=request.POST['owner_email'],
            contact=request.POST['contact'],
            morada=request.POST.get('morada', False) == 'on'
        )
    return redirect('home')

@login_required(login_url='login')
def adicionar_fornecedor(request):
    loja_nome = request.session.get('loja_atual')
    erro = ''
    sucesso = ''

    if request.method == 'POST':
        try:
            nif      = int(request.POST.get('nif'))
            nome     = request.POST.get('nome', '').strip()
            email    = request.POST.get('email', '').strip()
            contacto = int(request.POST.get('contacto'))
            morada   = request.POST.get('morada', '').strip()

            if Fornecedor.objects.filter(nif=nif).exists():
                erro = f'Já existe um fornecedor com o NIF {nif}.'
            elif not nome:
                erro = 'O nome não pode estar vazio.'
            else:
                Fornecedor.objects.create(
                    nif=nif, nome=nome, email=email,
                    contacto=contacto, morada=morada
                )
                sucesso = f'Fornecedor "{nome}" adicionado com sucesso!'
        except Exception as e:
            erro = f'Erro: {e}'

    return render(request, 'adicionar_fornecedor.html', {
        'loja_atual': loja_nome,
        'erro': erro,
        'sucesso': sucesso,
    })


@login_required(login_url='login')
def editar_produto(request, produto_id):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
            password = dados.get('password', '')
            user = authenticate(request, username=request.user.username, password=password)
            if not user:
                return JsonResponse({'erro': 'Password incorreta.'}, status=403)
            produto = get_object_or_404(Produto, id=produto_id)
            produto.preco      = float(dados['preco'])
            produto.desconto   = int(dados['desconto'])
            produto.iva        = int(dados['iva'])
            produto.qtd        = int(dados['qtd'])
            produto.seccao     = Seccao.objects.get(nome=dados['seccao'])
            fornecedor_nome    = dados.get('fornecedor')
            produto.fornecedor = Fornecedor.objects.get(nome=fornecedor_nome) if fornecedor_nome else None
            produto.save()
            return JsonResponse({'sucesso': True})
        except Exception as e:
            return JsonResponse({'erro': str(e)}, status=400)
    return JsonResponse({'erro': 'Método não permitido'}, status=405)

@login_required(login_url='login')
def ocultar_produto(request, produto_id):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
            password = dados.get('password', '')
            user = authenticate(request, username=request.user.username, password=password)
            if not user:
                return JsonResponse({'erro': 'Password incorreta.'}, status=403)
            produto = get_object_or_404(Produto, id=produto_id)
            produto.oculto = True
            produto.save()
            return JsonResponse({'sucesso': True})
        except Exception as e:
            return JsonResponse({'erro': str(e)}, status=400)
    return JsonResponse({'erro': 'Método não permitido'}, status=405)
@login_required(login_url='login')
def get_produto(request, produto_id):
    loja_nome = request.session.get('loja_atual')
    produto = get_object_or_404(Produto, id=produto_id)
    todas_seccoes = list(
        Seccao.objects.filter(loja__nomeloja=loja_nome).values_list('nome', flat=True)
    )
    todos_fornecedores = list(
        Fornecedor.objects.values_list('nome', flat=True).order_by('nome')
    )
    return JsonResponse({
        'id':                 produto.id,
        'nome':               produto.nome,
        'preco':              float(produto.preco),
        'qtd':                produto.qtd,
        'desconto':           produto.desconto if produto.desconto is not None else 0,
        'iva':                produto.iva if produto.iva is not None else 23,
        'seccao':             str(produto.seccao_id),
        'fornecedor':         produto.fornecedor.nome if produto.fornecedor else None,
        'todas_seccoes':      todas_seccoes,
        'todos_fornecedores': todos_fornecedores,
    })
